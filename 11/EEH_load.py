import pandas as pd
import numpy as np
import pyeeg
from numpy.random import randn
from pandas import ExcelWriter
from openpyxl import Workbook


def rle(inarray):
    """ run length encoding. Partial credit to R rle function. 
        Multi datatype arrays catered for including non Numpy
        returns: tuple (runlengths, startpositions, values) """
    ia = np.array(inarray)  # force numpy
    n = len(ia)
    if n == 0:
        return (None, None, None)
    else:
        y = np.array(ia[1:] != ia[:-1])  # pairwise unequal (string safe)
        i = np.append(np.where(y), n - 1)  # must include last element posi
        z = np.diff(np.append(-1, i))  # run lengths
        p = np.cumsum(np.append(0, z))[:-1]  # positions
        return (z, p, ia[i])


names1 = [u'Омельченко Сергей',
          u'Захаренков Антон',
          u'Ерин Федор',
          u'Гайнцева Татьяна',
          u'Антонова Нина']
names2 = [u'Едигарьев Иван',
          u'Дробот Олег',
          u'Брицын Евгений',
          u'Жибоедова Анастасия',
          u'Архипов Ярослав']
# folder = '2016-03-11_Collective_action'
state = 'Before_Soc'
label = 1
wb = Workbook()
wb.save(filename='{}_EEH.xlsx'.format(state))
for line in [1, 2]:
    for n in range(5):
        x = pd.read_csv('{}/{}/000{}_X'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'X'],
                        index_col=False)
        y = pd.read_csv('{}/{}/000{}_Y'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'Y'],
                        index_col=False)
        z = pd.read_csv('{}/{}/000{}_PZ'.format(state, line, 2 * n + 1), skiprows=10, sep=')', names=['time', 'Z'],
                        index_col=False)

        data = x.merge(y).merge(z)

        marker = pd.read_excel('{}/1/Markers_1-11.xlsx'.format(state))
        marker = marker[np.isfinite(marker['sec'])]
        data = data.iloc[:int(marker['tic-tot'].iloc[[-1]]) + 5000]

        df = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
        df = df[np.isfinite(df['X'])]
        df = df ** 2
        df = np.array(df.sum(axis=1))
        ndf = [df[i:i + 1500] for i in range(0, len(df), 1500)]

        mean = [j.mean() for j in ndf]
        std = [j.std() for j in ndf]
        energy = (np.array(mean) ** 2 + np.array(std)) ** (1. / 2)
        hurstl = [pyeeg.hurst(j) for j in ndf]

        de = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
        de = de[np.isfinite(de['X'])]
        nde = [np.array(de)[i:i + 1500] for i in range(0, len(df), 1500)]

        amax = [np.amax(j, axis=0) for j in nde]
        amin = [np.amin(j, axis=0) for j in nde]
        delta = (np.array(amax) - np.array(amin)) / 40
        nde = [(nde[j] - amin[j]) // delta[j] + 1 for j in range(len(amin))]
        t = [j[:, 0] * 1000000 + j[:, 1] * 1000 + j[:, 2] for j in nde]
        entropy = [-sum((rle(np.sort(i))[0] / 1500.) * np.log2(rle(np.sort(i))[0] / 1500.)) for i in t]

        from openpyxl import load_workbook

        book = load_workbook('{}_EEH.xlsx'.format(state))
        writer = pd.ExcelWriter('{}_EEH.xlsx'.format(state), engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        if (line == 1):
            data1 = pd.DataFrame(energy, columns=[names1[n]]).T.to_excel(writer, sheet_name=u'Energy',
                                                                         startrow=(line - 1) * 5 + n, startcol=1,
                                                                         header=False)
            data2 = pd.DataFrame(entropy, columns=[names1[n]]).T.to_excel(writer, sheet_name=u'Entropy',
                                                                          startrow=(line - 1) * 5 + n, startcol=1,
                                                                          header=False)
            data3 = pd.DataFrame(hurstl, columns=[names1[n]]).T.to_excel(writer, sheet_name=u'Hurst',
                                                                         startrow=(line - 1) * 5 + n, startcol=1,
                                                                         header=False)
        else:
            data1 = pd.DataFrame(energy, columns=[names2[n]]).T.to_excel(writer, sheet_name=u'Energy',
                                                                         startrow=(line - 1) * 5 + n, startcol=1,
                                                                         header=False)
            data2 = pd.DataFrame(entropy, columns=[names2[n]]).T.to_excel(writer, sheet_name=u'Entropy',
                                                                          startrow=(line - 1) * 5 + n, startcol=1,
                                                                          header=False)
            data3 = pd.DataFrame(hurstl, columns=[names2[n]]).T.to_excel(writer, sheet_name=u'Hurst',
                                                                         startrow=(line - 1) * 5 + n, startcol=1,
                                                                         header=False)

        writer.save()