import pandas as pd
import numpy as np
import pyeeg
from numpy.random import randn
from pandas import ExcelWriter
from openpyxl import Workbook


def rle(inarray):
    """ run length encoding. Partial credit to R rle function. 
        Multi datatype arrays catered for including non Numpy
        returns: tuple (runlengths, startpositions, values) """
    ia = np.array(inarray)  # force numpy
    n = len(ia)
    if n == 0:
        return (None, None, None)
    else:
        y = np.array(ia[1:] != ia[:-1])  # pairwise unequal (string safe)
        i = np.append(np.where(y), n - 1)  # must include last element posi
        z = np.diff(np.append(-1, i))  # run lengths
        p = np.cumsum(np.append(0, z))[:-1]  # positions
        return (z, p, ia[i])


names = [u'Преображенский Артём',
         u'Звагельский Роман',
         u'Коровина Екатерина',
         u'Шилова Алена',
         u'Михайлов Михаил',
         u'Гибадуллина Ольга',
         u'Пяточкин Михаил',
         u'Бредихина Анастасия',
         u'Маркашанская Дарья',
         u'Двинских Дарья']

# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 2
wb = Workbook()
wb.save(filename='{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        data = pd.read_excel("After_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))

        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()
        marker = pd.read_excel('After_Soc/1/Markers_1-2.xls')
        marker = marker[np.isfinite(marker['sec'])]

        lol1 = data.groupby('stad')
        c = [lol1.get_group(x) for x in lol1.groups]
        markers = [max(j.index) for j in c]
        markers.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        markers.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers = np.sort(markers)

        df = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
        df = df[np.isfinite(df['X'])]
        df = df ** 2
        df = np.array(df.sum(axis=1))
        ndf = []
        prev = 0

        for index in markers:
            if index > df.shape[0]: break
            ndf.append(np.array(df[prev:int(index)]))
            prev = int(index)

        if index < df.shape[0]: ndf.append(np.array(df[markers[-1]:]))

        mean = [j.mean() for j in ndf]
        std = [j.std() for j in ndf]
        energy = (np.array(mean) ** 2 + np.array(std)) ** (1. / 2)
        hurstl = [pyeeg.hurst(j) for j in ndf]

        de = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
        de = de[np.isfinite(de['X'])]
        nde = []
        prev = 0

        for index in markers:
            if index > de.shape[0]: break
            nde.append(np.array(de[prev:int(index)]))
            prev = int(index)

        if index < df.shape[0]: nde.append(np.array(de[markers[-1]:]))

        amax = [np.amax(j, axis=0) for j in nde]
        amin = [np.amin(j, axis=0) for j in nde]
        delta = (np.array(amax) - np.array(amin)) / 40
        nde = [(nde[j] - amin[j]) // delta[j] + 1 for j in range(len(amin))]
        t = [j[:, 0] * 1000000 + j[:, 1] * 1000 + j[:, 2] for j in nde]
        entropy = [-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i)))) for i in t]

        from openpyxl import load_workbook

        book = load_workbook('{}_EEH2.xlsx'.format(state))
        writer = pd.ExcelWriter('{}_EEH2.xlsx'.format(state), engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        data10 = pd.DataFrame(list(marker["Unnamed: 1"])).T.to_excel(writer, sheet_name=u'Energy', startrow=0,
                                                                     startcol=2, header=False)
        data20 = pd.DataFrame(list(marker["Unnamed: 1"])).T.to_excel(writer, sheet_name=u'Entropy', startrow=0,
                                                                     startcol=2, header=False)
        data30 = pd.DataFrame(list(marker["Unnamed: 1"])).T.to_excel(writer, sheet_name=u'Hurst', startrow=0,
                                                                     startcol=2, header=False)

        data1 = pd.DataFrame(energy, columns=[names[n]]).T.to_excel(writer, sheet_name=u'Energy', startrow=n + 1,
                                                                    startcol=1, header=False)
        data2 = pd.DataFrame(entropy, columns=[names[n]]).T.to_excel(writer, sheet_name=u'Entropy', startrow=n + 1,
                                                                     startcol=1, header=False)
        data3 = pd.DataFrame(hurstl, columns=[names[n]]).T.to_excel(writer, sheet_name=u'Hurst', startrow=n + 1,
                                                                    startcol=1, header=False)

        writer.save()