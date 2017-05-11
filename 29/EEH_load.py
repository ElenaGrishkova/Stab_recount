import pandas as pd
import numpy as np
import pyeeg
from numpy.random import randn
from pandas import ExcelWriter
from openpyxl import Workbook


def rle(inarray):
    """ run length encoding. Partial credit to R rle function. 
        Multi datatype arrays catered for including non Numpy
        returns: tuple (runlengths, startpositions, values) """
    ia = np.array(inarray)  # force numpy
    n = len(ia)
    if n == 0:
        return (None, None, None)
    else:
        y = np.array(ia[1:] != ia[:-1])  # pairwise unequal (string safe)
        i = np.append(np.where(y), n - 1)  # must include last element posi
        z = np.diff(np.append(-1, i))  # run lengths
        p = np.cumsum(np.append(0, z))[:-1]  # positions
        return (z, p, ia[i])


names = [u'Беликова Мария',
         u'Михайлов Владислав',
         u'Анисимов Александр',
         u'Яновская Елизавета',
         u'Плотников Аким',
         u'Бальков Андрей',
         u'Дмитриев Егор',
         u'Полукеева Вера',
         u'Ракутин Юрий',
         u'Слижикова Анна']

# folder = '2016-03-11_Collective_action'
state = 'Before_Soc'

wb = Workbook()
wb.save(filename='{}_EEH.xlsx'.format(state))
for line in [1]:
    for n in range(5, 10):

        if (2 * n < 9): x = pd.read_csv('{}/1/000{}_X'.format(state, 2 * n), skiprows=10, sep=')', names=['time', 'X'],
                                        index_col=False)
        if (2 * n < 9): y = pd.read_csv('{}/1/000{}_Y'.format(state, 2 * n), skiprows=10, sep=')', names=['time', 'Y'],
                                        index_col=False)
        if (2 * n < 9): z = pd.read_csv('{}/1/000{}_PZ'.format(state, 2 * n + 1), skiprows=10, sep=')',
                                        names=['time', 'Z'], index_col=False)

        if (2 * n > 9): x = pd.read_csv('{}/2/000{}_X'.format(state, 2 * (n - 5)), skiprows=10, sep=')',
                                        names=['time', 'X'], index_col=False)
        if (2 * n > 9): y = pd.read_csv('{}/2/000{}_Y'.format(state, 2 * (n - 5)), skiprows=10, sep=')',
                                        names=['time', 'Y'], index_col=False)
        if (2 * n > 9): z = pd.read_csv('{}/2/000{}_PZ'.format(state, 2 * (n - 5) + 1), skiprows=10, sep=')',
                                        names=['time', 'Z'], index_col=False)

        # info = pd.read_excel('160401_SO1-2.xlsx',sheetname=u'Г1', index_col=0)
        # info2 = pd.read_excel('160401_SO1-2.xlsx',sheetname=u'Г2', index_col=0)
        data = x.merge(y).merge(z)
        # data['marker'] = '0'

        # zt = pd.read_csv('160401_1555.csv',skiprows=32, sep = ';')
        # a = list(zt.columns)
        # a[0] = 'lol'
        # zt.columns = a
        # start = float(np.array(zt[zt.subjects == 'subjects'][zt.lol == 3][zt.Period == '1'].iloc[[1],[9]])[0][0])
        # start2 = float(np.array(zt[zt.subjects == 'subjects'][zt.lol == 4][zt.Period == '1'].iloc[[1],[9]])[0][0])


        marker = pd.read_excel('Before_Soc/2/Markers_2-1.xls')
        marker = marker[np.isfinite(marker['sec'])]
        start_extract = float(np.array(marker.iloc[[2], [3]])[0][0])
        start_contrib = float(np.array(marker.iloc[[9], [3]])[0][0])

        # data = data.iloc[[j for j in range (0,np.array(marker['tic-tot'].astype(int))[22])]]

        df = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
        df = df[np.isfinite(df['X'])]
        df = df ** 2
        df = np.array(df.sum(axis=1))
        ndf = [df[i:i + 1500] for i in range(0, len(df), 1500)]

        mean = [j.mean() for j in ndf]
        std = [j.std() for j in ndf]
        energy = (np.array(mean) ** 2 + np.array(std)) ** (1. / 2)
        hurstl = [pyeeg.hurst(j) for j in ndf]

        de = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
        de = de[np.isfinite(de['X'])]
        nde = [np.array(de)[i:i + 1500] for i in range(0, len(df), 1500)]

        amax = [np.amax(j, axis=0) for j in nde]
        amin = [np.amin(j, axis=0) for j in nde]
        delta = (np.array(amax) - np.array(amin)) / 40
        nde = [(nde[j] - amin[j]) // delta[j] + 1 for j in range(len(amin))]
        t = [j[:, 0] * 1000000 + j[:, 1] * 1000 + j[:, 2] for j in nde]
        entropy = [-sum((rle(np.sort(i))[0] / 1500.) * np.log2(rle(np.sort(i))[0] / 1500.)) for i in t]

        from openpyxl import load_workbook

        book = load_workbook('{}_EEH.xlsx'.format(state))
        writer = pd.ExcelWriter('{}_EEH.xlsx'.format(state), engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        data1 = pd.DataFrame(energy, columns=[names[n]]).T.to_excel(writer, sheet_name=u'Energy', startrow=n,
                                                                    startcol=1, header=False)
        data2 = pd.DataFrame(entropy, columns=[names[n]]).T.to_excel(writer, sheet_name=u'Entropy', startrow=n,
                                                                     startcol=1, header=False)
        data3 = pd.DataFrame(hurstl, columns=[names[n]]).T.to_excel(writer, sheet_name=u'Hurst', startrow=n, startcol=1,
                                                                    header=False)

        writer.save()