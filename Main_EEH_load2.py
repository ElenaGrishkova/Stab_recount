# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import pyeeg
import logging
from openpyxl import Workbook

INFO_COL_TYPES={'time':unicode,'X':float,'Y':float,'Z':float,'stad':unicode,
                'Profit':int,'TP':int,'LP':int,'Sum':int,'TH':float}

np.seterr(divide='print', invalid='ignore')

logging.basicConfig(filename='log.txt',level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.debug('Start')

def rle(inarray):
    """ run length encoding. Partial credit to R rle function. 
        Multi datatype arrays catered for including non Numpy
        returns: tuple (runlengths, startpositions, values) """
    ia = np.array(inarray)  # force numpy
    n = len(ia)
    if n == 0:
        return (None, None, None)
    else:
        y = np.array(ia[1:] != ia[:-1])  # pairwise unequal (string safe)
        i = np.append(np.where(y), n - 1)  # must include last element posi
        z = np.diff(np.append(-1, i))  # run lengths
        p = np.cumsum(np.append(0, z))[:-1]  # positions
        return (z, p, ia[i])

def getDF_V2(data):
    # V**2
    df = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['X'])]
    df = df ** 2
    df = np.array(df.sum(axis=1))
    return df
def getDF_Vx2_Vy2(data):
    # Vx**2+Vy**2
    df = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['X'])]
    df = df ** 2
    df = np.array(df['X']+df['Y'])
    return df
def getDF_Vx2(data):
    # Vx**2
    df = data[['X']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['X'])]
    df = df ** 2
    df = np.array(df)
    return df
def getDF_Vy2(data):
    # Vy**2
    df = data[['Y']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['Y'])]
    df = df ** 2
    df = np.array(df)
    return df
def getDF_Vz2(data):
    # Vz**2
    df = data[['Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['Z'])]
    df = df ** 2
    df = np.array(df)
    return df
def getDF_V(data):
    # V
    df = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['X'])]
    df = df ** 2
    df = np.array(df.sum(axis=1))
    return df ** (1. / 2)
def getDF_Vx(data):
    # Vx
    df = data[['X']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['X'])]
    df = np.array(df)
    return df
def getDF_Vy(data):
    # Vy
    df = data[['Y']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['Y'])]
    df = np.array(df)
    return df
def getDF_Vz(data):
    # Vz
    df = data[['Z']].rolling(window=3, min_periods=1, center=True).mean().diff()
    df = df[np.isfinite(df['Z'])]
    df = np.array(df)
    return df

def calcNDF(df, markers, isComment) :
    ndf = []
    prev = 0
    for index in markers:
        if index > df.shape[0]: break
        ndf.append(np.array(df[prev:int(index)]))
        prev = int(index)
    if not isComment:
        if index < df.shape[0]:
            ndf.append(np.array(df[int(markers[-1]):]))
    return df,  ndf

def processData(data, line, n, isComment, folder, markers, marker1, isNormal=False) :
    ndf_all={}
    ndf_all['V2']=calcNDF(getDF_V2(data), markers, isComment)
    ndf_all['Vx2+Vy2']=calcNDF(getDF_Vx2_Vy2(data), markers, isComment)
    ndf_all['Vx2']=calcNDF(getDF_Vx2(data), markers, isComment)
    ndf_all['Vy2']=calcNDF(getDF_Vy2(data), markers, isComment)
    ndf_all['Vz2']=calcNDF(getDF_Vz2(data), markers, isComment)
    ndf_all['Vx'] = calcNDF(getDF_Vx(data), markers, isComment)
    ndf_all['Vy'] = calcNDF(getDF_Vy(data), markers, isComment)
    ndf_all['Vz'] = calcNDF(getDF_Vz(data), markers, isComment)

    energy={}
    for key, (df, ndf) in ndf_all.iteritems() :
        mean = [j.mean() for j in ndf]
        std = [j.std() for j in ndf]
        energy[key] = np.log((np.array(mean) ** 2 + np.array(std) ** 2) ** (1. / 2))

    hurst={}
    for key, (df, ndf) in ndf_all.iteritems():
        hurst[key] = [pyeeg.hurst(j) for j in ndf]

#Внимание! Для энтропии убрала diff, чтобы считал не по скоростям, а по самим точкам,
    # но нужно ли оставлять усреднение по трем точкам? Сейчас оно оставлено
    de = data[['X', 'Y', 'Z']].rolling(window=3, min_periods=1, center=True).mean()
    de = de[np.isfinite(de['X'])]
    nde = []
    prev = 0

    for index in markers:
        if index > de.shape[0]: break
        nde.append(np.array(de[prev:int(index)]))
        prev = int(index)

    if not isComment:
        if index<df.shape[0]: nde.append(np.array(de[int(markers[-1]):]))

    amax = [np.amax(j, axis=0) for j in nde]
    amin = [np.amin(j, axis=0) for j in nde]

    entropy={}
    delta1 = (np.array(amax) - np.array(amin)) / 40
    nde1 = [(nde[j] - amin[j]) // delta1[j] + 1 for j in range(len(amin))]
    t1 = [j[:, 0] * 1000000 + j[:, 1] * 1000 + j[:, 2] for j in nde1]
    entropy['old'] = [(-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i))))) / np.log2(1600) for i in t1]

    tX = [j[:, 0] * 1000000 for j in nde1]
    entropy['X old'] = [(-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i))))) / np.log2(1600) for i in tX]

    tY = [j[:, 1] * 1000 for j in nde1]
    entropy['Y old'] = [(-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i))))) / np.log2(1600) for i in tY]

    tZ = [j[:, 2] for j in nde1]
    entropy['Z old'] = [(-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i))))) / np.log2(1600) for i in tZ]

    delta2 = [(amax[j] - amin[j]) / (len(nde[j])) ** (1./2) for j in range(len(amin))]
    nde2 = [(nde[j] - amin[j]) // delta2[j] + 1 for j in range(len(amin))]
    t2 = [j[:, 0] * 1000000 + j[:, 1] * 1000 + j[:, 2] for j in nde2]
    entropy['new'] = [(-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i))))) / np.log2(float(len(i))) for i in t2]

    tX = [j[:, 0] * 1000000 for j in nde2]
    entropy['X'] = [(-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i))))) / np.log2(float(len(i))) for i in tX]

    tY = [j[:, 1] * 1000 for j in nde2]
    entropy['Y'] = [(-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i))))) / np.log2(float(len(i))) for i in tY]

    tZ = [j[:, 2] for j in nde2]
    entropy['Z'] = [(-sum((rle(np.sort(i))[0] / float(len(i))) * np.log2(rle(np.sort(i))[0] / float(len(i))))) / np.log2(float(len(i))) for i in tZ]

    from openpyxl import load_workbook

    book = load_workbook(folder+'/{}_EEH2.xlsx'.format(state))
    writer = pd.ExcelWriter(folder+'/{}_EEH2.xlsx'.format(state), engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    for key in energy.iterkeys():
        pd.DataFrame(marker1).T.to_excel(writer, sheet_name=u'Energy '+unicode(key), startrow=0, startcol=1, header=False)
    for key in entropy.iterkeys():
        pd.DataFrame(marker1).T.to_excel(writer, sheet_name=u'Entropy '+unicode(key), startrow=0, startcol=1, header=False)
    for key in energy.iterkeys():
        pd.DataFrame(marker1).T.to_excel(writer, sheet_name=u'Hurst '+unicode(key), startrow=0, startcol=1, header=False)

    columns=[names[n] if line == 1 else names2[n]]
    for key, energy_item in energy.iteritems() :
        pd.DataFrame(energy_item, columns=columns).T.to_excel(writer, sheet_name=u'Energy '+unicode(key),
                                                              startrow=(line - 1) * 5 + n + 1,
                                                                                             startcol=1, header=False)
    for key, entropy_item in entropy.iteritems():
        pd.DataFrame(entropy_item, columns=columns).T.to_excel(writer, sheet_name=u'Entropy '+unicode(key),
                                                           startrow=(line - 1) * 5 + n + 1,
                                                                                     startcol=1, header=False)

    for key, hurst_item in hurst.iteritems():
        pd.DataFrame(hurst_item, columns=columns).T.to_excel(writer, sheet_name=u'Hurst '+unicode(key),
                                                                                            startrow=(line - 1) * 5 + n + 1,
                                                                                            startcol=1, header=False)

    writer.save()


#-----------------------------------------------------------------------------------------------------------------------
#01 Before
#-----------------------------------------------------------------------------------------------------------------------
logging.debug('01 Before')
names = [u'Иванов Владислав',
         u'Леонтьева Ксения',
         u'Кравец Мария',
         u'Тукина Юлия',
         u'Дубенко Антон',
         u'Акулов Дмитрий',
         u'Гаврильчик Богдан',
         u'Котова Маргарита',
         u'Гуров Алексей',
         u'Федотов Никита']
names2=[]

# folder = '2016-03-11_Collective_action'
state = 'Before_Soc'
label = 2
wb = Workbook()
wb.save(filename='01/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        if n == 1: continue
        #data = pd.read_excel("01/Before_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"01/Before_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)

        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()
        marker = pd.read_excel('01/Before_Soc/1/Markers_1-1.xls')
        marker = marker[np.isfinite(marker['sec'])]

        lol1 = data.groupby('stad')
        c = [lol1.get_group(x) for x in lol1.groups]
        markers = [max(j.index) for j in c]
        markers.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        markers.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers = np.sort(markers)

        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', 
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'ГО', u'ГЗ']

        processData(data, line, n, False, '01', markers, marker1)

#-----------------------------------------------------------------------------------------------------------------------
#01 After
#-----------------------------------------------------------------------------------------------------------------------

logging.debug('01 After')
names = [u'Гуров Алексей',
         u'Калмыкова Светлана',
         u'Акулов Дмитрий',
         u'Гаврильчик Богдан',
         u'Котова Маргарита',
         u'Прокопов Иван',
         u'Дубенко Антон',
         u'Тукина Юлия',
         u'Федотов Никита',
         u'Кравец Мария']

# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 2
wb = Workbook()
wb.save(filename='01/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("01/After_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"01/After_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)

        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()
        marker = pd.read_excel('01/After_Soc/1/Markers_1-2.xls')
        marker = marker[np.isfinite(marker['sec'])]

        lol1 = data.groupby('stad')
        c = [lol1.get_group(x) for x in lol1.groups]
        markers = [max(j.index) for j in c]
        #markers.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        #markers.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers = np.sort(markers)

        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10']
        processData(data, line, n, False, '01', markers, marker1)


#-----------------------------------------------------------------------------------------------------------------------
#08 Before
#-----------------------------------------------------------------------------------------------------------------------
logging.debug('08 Before')
names = [u'Самолыго Алексей',
         u'Амбарцумян Дарья',
         u'Заруба Кирилл',
         u'Рыжников Андрей',
         u'Гребенчук Сергей',
         u'Казаков Алексей',
         u'Наркунас Татьяна',
         u'Замятина Екатерина',
         u'Ермолова Марина',
         u'Лелякин Дмитрий']

# folder = '2016-03-11_Collective_action'
state = 'Before_Soc'
label = 2
wb = Workbook()
wb.save(filename='08/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("08/Before_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"08/Before_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)
        data1 = data[:41787]
        data2 = data[41788:]
        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()

        marker = pd.read_excel('08/Before_Soc/1/Markers_1-1.xls')
        marker = marker[np.isfinite(marker['sec'])]

        data1 = data1.drop(data1[data1.stad == '0'].index)
        data1 = data1.reset_index()
        data2 = data2.drop(data2[data2.stad == '0'].index)
        data2 = data2.reset_index()

        lol1 = data1.groupby('stad')
        c1 = [lol1.get_group(x) for x in lol1.groups]
        markers1 = [max(j.index) for j in c1]
        markers1.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        markers1.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers1 = np.sort(markers1)

        lol2 = data2.groupby('stad')
        c2 = [lol2.get_group(x) for x in lol2.groups]
        markers2 = [max(j.index) for j in c2]
        markers2 = np.sort(markers2)

        markers = list(markers1) + list(markers2 + markers1[-1])

        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5',
        u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'ГО', u'ГЗ']
        processData(data, line, n, True, '08', markers, marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 08 After
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('08 After')
names = [u'Садакова Кристина',
         u'Наркунас Татьяна',
         u'Лелякин Дмитрий',
         u'Рыжников Андрей',
         u'Гребенчук Сергей',
         u'Казаков Алексей',
         u'Заруба Кирилл',
         u'Замятина Екатерина',
         u'Ермолова Марина',
         u'Самолыго Алексей']

# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 2
wb = Workbook()
wb.save(filename='08/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("08/After_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"08/After_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)
        data1 = data[:28012]
        data2 = data[28012:]
        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()

        marker = pd.read_excel('08/After_Soc/1/Markers_1-2.xls')
        marker = marker[np.isfinite(marker['sec'])]

        data1 = data1.drop(data1[data1.stad == '0'].index)
        data1 = data1.reset_index()
        data2 = data2.drop(data2[data2.stad == '0'].index)
        data2 = data2.reset_index()

        lol1 = data1.groupby('stad')
        c1 = [lol1.get_group(x) for x in lol1.groups]
        markers1 = [max(j.index) for j in c1]
        markers1.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        markers1.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers1 = np.sort(markers1)

        lol2 = data2.groupby('stad')
        c2 = [lol2.get_group(x) for x in lol2.groups]
        markers2 = [max(j.index) for j in c2]
        markers2 = np.sort(markers2)

        markers = list(markers1) + list(markers2 + markers1[-1])

        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5',
        u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'ГО', u'ГЗ']
        processData(data, line, n, True, '08', markers, marker1)


# -----------------------------------------------------------------------------------------------------------------------
# 11 Before
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('11 Before')

names = [u'Омельченко Сергей',
          u'Захаренков Антон',
          u'Ерин Федор',
          u'Гайнцева Татьяна',
          u'Антонова Нина']
names2 = [u'Едигарьев Иван',
          u'Дробот Олег',
          u'Брицын Евгений',
          u'Жибоедова Анастасия',
          u'Архипов Ярослав']
# folder = '2016-03-11_Collective_action'
state = 'Before_Soc'
label = 1
wb = Workbook()
wb.save(filename='11/{}_EEH2.xlsx'.format(state))
for line in [1, 2]:
    for n in range(5):
        logging.debug('n: %d', n)
        x = pd.read_csv('11/{}/{}/000{}_X'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'X'],
                        index_col=False)
        y = pd.read_csv('11/{}/{}/000{}_Y'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'Y'],
                        index_col=False)
        z = pd.read_csv('11/{}/{}/000{}_PZ'.format(state, line, 2 * n + 1), skiprows=10, sep=')', names=['time', 'Z'],
                        index_col=False)

        data = x.merge(y).merge(z)

        marker = pd.read_excel('11/{}/1/Markers_1-11.xlsx'.format(state))
        marker = marker[np.isfinite(marker['sec'])]
        data = data.iloc[:int(marker['tic-tot'].iloc[[-1]]) + 5000]

        marker1 = [u'0', u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'10',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'10', u'11', u'Extr_end', u'ГО', u'ГЗ', u'ГЗ_end']
        processData(data, line, n, False, '11', np.array(marker["tic-tot"]), marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 11 After
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('11 After')

names = [u'Дробот Олег',
          u'Жибоедова Анастасия',
          u'Едигарьев Иван',
          u'Брицын Евгений',
          u'Антонова Нина']
names2 = [u'Захаренков Антон',
          u'Ерин Федор',
          u'Омельченко Сергей',
          u'Гайнцева Татьяна',
          u'Архипов Ярослав']
# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 1
wb = Workbook()
wb.save(filename='11/{}_EEH2.xlsx'.format(state))
for line in [1, 2]:
    for n in range(5):
        logging.debug('n: %d', n)
        x = pd.read_csv('11/{}/{}/000{}_X'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'X'],
                        index_col=False)
        y = pd.read_csv('11/{}/{}/000{}_Y'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'Y'],
                        index_col=False)
        z = pd.read_csv('11/{}/{}/000{}_PZ'.format(state, line, 2 * n + 1), skiprows=10, sep=')', names=['time', 'Z'],
                        index_col=False)

        data = x.merge(y).merge(z)


        marker = pd.read_excel('11/{}/1/Markers_1-2.xls'.format(state))
        marker = marker[np.isfinite(marker['sec'])]
        data = data.iloc[:int(marker['tic-tot'].iloc[[-1]]) + 5000]

        marker1 = [u'0', u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'10',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'10', u'Extr_end', u'ГО', u'ГЗ', u'ГЗ_end']
        processData(data, line, n, False, '11', np.array(marker["tic-tot"]), marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 15 Before
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('15 Before')
names = [u'Григорик Николай',
        u'Юдин Александр',
        u'Макаровский Дмитрий',
        u'Михальченко Егор',
        u'Артемова Надежда',
        u'Рассолов Сергей',
        u'Наседкин Илья',
        u'Панин Артем',
        u'Камчаткин Владимир',
        u'Гайнуллин Дмитрий']
state = 'Before_Soc'
wb = Workbook()
wb.save(filename='15/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        #data = pd.read_excel("15/Before_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"15/Before_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)
        data1 = data[:37316]
        data2 = data[37316:]
        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()

        marker = pd.read_excel('15/After_Soc/2/Markers_2-2.xls')
        marker = marker[np.isfinite(marker['sec'])]

        data1 = data1.drop(data1[data1.stad == '0'].index)
        data1 = data1.reset_index()
        data2 = data2.drop(data2[data2.stad == '0'].index)
        data2 = data2.reset_index()

        lol1 = data1.groupby('stad')
        c1 = [lol1.get_group(x) for x in lol1.groups]
        markers1 = [max(j.index) for j in c1]
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers1 = np.sort(markers1)

        lol2 = data2.groupby('stad')
        c2 = [lol2.get_group(x) for x in lol2.groups]
        markers2 = [max(j.index) for j in c2]
        markers2 = np.sort(markers2)

        markers = list(markers1) + list(markers2 + markers1[-1])

        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5']
        processData(data, line, n, True, '15', markers, marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 15 After
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('15 After')

names = [u'Гайнуллин Дмитрий',
         u'Юдин Александр',
         u'Макаровский Дмитрий',
         u'Михальченко Егор',
         u'Рассолов Сергей',
         u'Пашкова Мария',
         u'Наседкин Илья',
         u'Панин Артем',
         u'Камчаткин Владимир',
         u'Григорик Николай']

# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 2
wb = Workbook()
wb.save(filename='15/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(5, 10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("15/After_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"15/After_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)
        data1 = data[:23368]
        data2 = data[23368:]
        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()

        marker = pd.read_excel('15/After_Soc/2/Markers_2-2.xls')
        marker = marker[np.isfinite(marker['sec'])]

        data1 = data1.drop(data1[data1.stad == '0'].index)
        data1 = data1.reset_index()
        data2 = data2.drop(data2[data2.stad == '0'].index)
        data2 = data2.reset_index()

        lol1 = data1.groupby('stad')
        c1 = [lol1.get_group(x) for x in lol1.groups]
        markers1 = [max(j.index) for j in c1]
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers1 = np.sort(markers1)

        lol2 = data2.groupby('stad')
        c2 = [lol2.get_group(x) for x in lol2.groups]
        markers2 = [max(j.index) for j in c2]
        markers2 = np.sort(markers2)

        markers = list(markers1) + list(markers2 + markers1[-1])
        marker1 = [u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5']
        processData(data, line, n, True, '15', markers, marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 18 Before
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('18 Before')

names = [u'Николаева Анна',
u'Лескин Иван',
u'Магнитов Михаил',
u'Пономарев Евгений',
u'Ишханян Аршак']
names2 = [
u'Герасимов Антонн',
u'Рулев Георгий',
u'Биктайров Роман',
u'Шилов Иннокентий',
u'Мухаметдинов Гаяз']

# folder = '2016-03-11_Collective_action'
state = 'Before_Soc'
label = 1
wb = Workbook()
wb.save(filename='18/{}_EEH2.xlsx'.format(state))
for line in [1, 2]:
    for n in range(5):
        logging.debug('n: %d', n)
        x = pd.read_csv('18/{}/{}/000{}_X'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'X'],
                        index_col=False)
        y = pd.read_csv('18/{}/{}/000{}_Y'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'Y'],
                        index_col=False)
        z = pd.read_csv('18/{}/{}/000{}_PZ'.format(state, line, 2 * n + 1), skiprows=10, sep=')', names=['time', 'Z'],
                        index_col=False)

        data = x.merge(y).merge(z)

        marker = pd.read_excel('18/{}/1/Markers_1-{}.xls'.format(state, label))
        marker = marker[np.isfinite(marker['sec'])]
        data = data.iloc[:int(marker['tic-tot'].iloc[[-1]]) + 5000]

        marker1 = [u'0', u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'10', u'Extr_end', u'ГО', u'ГЗ', u'ГЗ_end']
        processData(data, line, n, False, '18', np.array(marker["tic-tot"]), marker1)
# -----------------------------------------------------------------------------------------------------------------------
# 18 After
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('18 After')

names = [u'Николаева Анна',
          u'Мухаметдинов Гаяз',
          u'Магнитов Михаил',
          u'Пономарев Евгений',
          u'Ишханян Аршак']
names2 = [
    u'Исрафилов Ильдар',
    u'Рулев Георгий',
    u'Биктайров Роман',
    u'Шилов Иннокентий',
    u'Лескин Иван']
# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 2
wb = Workbook()
wb.save(filename='18/{}_EEH2.xlsx'.format(state))
for line in [1, 2]:
    for n in range(5):
        logging.debug('n: %d', n)
        if line == 1:
            x = pd.read_csv('18/{}/{}/000{}_X'.format(state, line, 3 * n), skiprows=10, sep=')', names=['time', 'X'],
                            index_col=False)
        else:
            x = pd.read_csv('18/{}/{}/000{}_X'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'X'],
                            index_col=False)

        if line == 1:
            y = pd.read_csv('18/{}/{}/000{}_Y'.format(state, line, 3 * n), skiprows=10, sep=')', names=['time', 'Y'],
                            index_col=False)
        else:
            y = pd.read_csv('18/{}/{}/000{}_Y'.format(state, line, 2 * n), skiprows=10, sep=')', names=['time', 'Y'],
                            index_col=False)
        if line == 1: z = pd.read_csv('18/{}/{}/000{}_Z'.format(state, line, 3 * n + 1), skiprows=10, sep=')',
                                      names=['time', 'Z'], index_col=False)
        if line == 2: z = pd.read_csv('18/{}/{}/000{}_PZ'.format(state, line, 2 * n + 1), skiprows=10, sep=')',
                                      names=['time', 'Z'], index_col=False)

        data = x.merge(y).merge(z)

        marker = pd.read_excel('18/{}/1/Markers_1-{}.xls'.format(state, label))
        marker = marker[np.isfinite(marker['sec'])]
        data = data.iloc[:int(marker['tic-tot'].iloc[[-1]]) + 5000]
        marker1 = [u'0', u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'10',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'10', u'Extr_end']
        processData(data, line, n, False, '18', np.array(marker["tic-tot"]), marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 22 Before
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('22 Before')

names = [u'Ильинская Олеся',
         u'Ностаева Арина',
         u'Елизарова Анна',
         u'Латыпова Галия',
         u'Никулина Ирина',
         u'Левин Евгений',
         u'Староватых Юлия',
         u'Столяренко Мария',
         u'Якимчиков Максим',
         u'Кузнецов Никита']

# folder = '2016-03-11_Collective_action'
state = 'Before_Soc'

wb = Workbook()
wb.save(filename='22/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("22/Before_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"22/Before_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)
        data1 = data[:41235]
        data2 = data[41235:]
        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()

        marker = pd.read_excel('22/After_Soc/1/Markers_1-2.xls')
        marker = marker[np.isfinite(marker['sec'])]

        data1 = data1.drop(data1[data1.stad == '0'].index)
        data1 = data1.reset_index()
        data2 = data2.drop(data2[data2.stad == '0'].index)
        data2 = data2.reset_index()

        lol1 = data1.groupby('stad')
        c1 = [lol1.get_group(x) for x in lol1.groups]
        markers1 = [max(j.index) for j in c1]
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers1 = np.sort(markers1)

        lol2 = data2.groupby('stad')
        c2 = [lol2.get_group(x) for x in lol2.groups]
        markers2 = [max(j.index) for j in c2]
        markers2 = np.sort(markers2)

        markers = list(markers1) + list(markers2 + markers1[-1])

        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'ГО', u'ГЗ']
        processData(data, line, n, True, '22', markers, marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 22 After
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('22 After')


names = [u'Кузнецов Никита',
         u'Ильинская Олеся',
         u'Никулина Ирина',
         u'Волков Александр',
         u'Столяренко Мария',
         u'Ностаева Арина',
         u'Староватых Юлия',
         u'Латыпова Галия',
         u'Елизарова Анна',
         u'Левин Евгений']

# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 2
wb = Workbook()
wb.save(filename='22/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("22/After_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"22/After_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)
        data1 = data[:26704]
        data2 = data[26704:]
        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()

        marker = pd.read_excel('22/After_Soc/1/Markers_1-2.xls')
        marker = marker[np.isfinite(marker['sec'])]

        data1 = data1.drop(data1[data1.stad == '0'].index)
        data1 = data1.reset_index()
        data2 = data2.drop(data2[data2.stad == '0'].index)
        data2 = data2.reset_index()

        lol1 = data1.groupby('stad')
        c1 = [lol1.get_group(x) for x in lol1.groups]
        markers1 = [max(j.index) for j in c1]
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers1 = np.sort(markers1)

        lol2 = data2.groupby('stad')
        c2 = [lol2.get_group(x) for x in lol2.groups]
        markers2 = [max(j.index) for j in c2]
        markers2 = np.sort(markers2)

        markers = list(markers1) + list(markers2 + markers1[-1])
        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5',
        u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'ГО', u'ГЗ']
        processData(data, line, n, True, '22', markers, marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 25 Before
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('25 Before')
names=[
u'Преображенский Артём',
u'Маркашанская Дарья',
u'Коровина Екатерина',
u'Князев Данила',
u'Михайлов Михаил',
u'Харламов Дмитрий',
u'Пяточкин Михаил',
u'Бредихина Анастасия',
u'Двинских Дарья',
u'Шилова Алена',
]
state = 'Before_Soc'
label = 2
wb = Workbook()
wb.save(filename='25/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("25/Before_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"25/Before_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)

        data = data.drop(data[data.stad == '0'].index)

        data = data.reset_index()
        marker = pd.read_excel('25/Before_Soc/1/Markers_1-1.xls')
        marker = marker[np.isfinite(marker['sec'])]

        lol12 = data.groupby('stad')
        c = [lol12.get_group(x) for x in lol12.groups]
        markers = [max(j.index) for j in c]
        #markers.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        #markers.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers = np.sort(markers)
        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10',
                   u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10']
        processData(data, line, n, False, '25', markers, marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 25 After
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('25 After')

names = [u'Преображенский Артём',
         u'Звагельский Роман',
         u'Коровина Екатерина',
         u'Шилова Алена',
         u'Михайлов Михаил',
         u'Гибадуллина Ольга',
         u'Пяточкин Михаил',
         u'Бредихина Анастасия',
         u'Маркашанская Дарья',
         u'Двинских Дарья']

# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 2
wb = Workbook()
wb.save(filename='25/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("25/After_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"25/After_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)

        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()
        marker = pd.read_excel('25/After_Soc/1/Markers_1-2.xls')
        marker = marker[np.isfinite(marker['sec'])]

        lol1 = data.groupby('stad')
        c = [lol1.get_group(x) for x in lol1.groups]
        markers = [max(j.index) for j in c]
        markers.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        markers.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers = np.sort(markers)
        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10',
                   u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'6', u'6', u'7', u'7', u'8', u'8', u'9', u'9', u'10', u'ГО', u'ГЗ']
        processData(data, line, n, False, '25', markers, marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 29 Before
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('29 Before')

names = [u'Беликова Мария',
         u'Михайлов Владислав',
         u'Анисимов Александр',
         u'Яновская Елизавета',
         u'Плотников Аким',
         u'Бальков Андрей',
         u'Дмитриев Егор',
         u'Полукеева Вера',
         u'Ракутин Юрий',
         u'Слижикова Анна']

# folder = '2016-03-11_Collective_action'
state = 'Before_Soc'
label = 2
wb = Workbook()
wb.save(filename='29/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(5, 10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("29/Before_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"29/Before_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)
        data1 = data[:33901]
        data2 = data[33901:]
        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()

        marker = pd.read_excel('29/Before_Soc/2/Markers_2-1.xls')
        marker = marker[np.isfinite(marker['sec'])]

        data1 = data1.drop(data1[data1.stad == '0'].index)
        data1 = data1.reset_index()
        data2 = data2.drop(data2[data2.stad == '0'].index)
        data2 = data2.reset_index()

        lol1 = data1.groupby('stad')
        c1 = [lol1.get_group(x) for x in lol1.groups]
        markers1 = [max(j.index) for j in c1]
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers1 = np.sort(markers1)

        lol2 = data2.groupby('stad')
        c2 = [lol2.get_group(x) for x in lol2.groups]
        markers2 = [max(j.index) for j in c2]
        markers2 = np.sort(markers2)

        markers = list(markers1) + list(markers2 + markers1[-1])
        marker1 = [u'ГО', u'ГЗ', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5',
                  u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5']
        processData(data, line, n, True, '29', markers, marker1)

# -----------------------------------------------------------------------------------------------------------------------
# 29 After
# -----------------------------------------------------------------------------------------------------------------------
logging.debug('29 After')

names = [u'Беликова Мария',
u'Ракутин Юрий',
u'Анисимов Александр',
u'Полукеева Вера',
u'Плотников Аким',
u'Слижикова Анна',
u'Бальков Андрей',
u'Яновская Елизавета',
u'Дмитриев Егор',
u'Михайлов Владислав']

# folder = '2016-03-11_Collective_action'
state = 'After_Soc'
label = 2
wb = Workbook()
wb.save(filename='29/{}_EEH2.xlsx'.format(state))
for line in [1]:
    for n in range(10):
        logging.debug('n: %d', n)
        #data = pd.read_excel("29/After_Soc_INFO.xlsx", sheetname=u'{}{}'.format(n + 1, names[n]))
        data = pd.read_csv(u"29/After_Soc_INFO "+u'{}{}'.format(n + 1, names[n])+u".csv", dtype=INFO_COL_TYPES)
        data1 = data[:18971]
        data2 = data[18971:]
        data = data.drop(data[data.stad == '0'].index)
        data = data.reset_index()

        marker = pd.read_excel('29/Before_Soc/2/Markers_2-1.xls')
        marker = marker[np.isfinite(marker['sec'])]

        data1 = data1.drop(data1[data1.stad == '0'].index)
        data1 = data1.reset_index()
        data2 = data2.drop(data2[data2.stad == '0'].index)
        data2 = data2.reset_index()

        lol1 = data1.groupby('stad')
        c1 = [lol1.get_group(x) for x in lol1.groups]
        markers1 = [max(j.index) for j in c1]
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[0]])[0]))
        # markers1.append(int(np.array(marker['tic-tot'].iloc[[1]])[0]))
        markers1 = np.sort(markers1)

        lol2 = data2.groupby('stad')
        c2 = [lol2.get_group(x) for x in lol2.groups]
        markers2 = [max(j.index) for j in c2]
        markers2 = np.sort(markers2)

        markers = list(markers1) + list(markers2 + markers1[-1])
        marker1 = [u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4',u'4', u'5',
                  u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'5', u'1', u'1', u'2', u'2', u'3', u'3', u'4', u'4', u'5', u'ГО', u'ГЗ']
        processData(data, line, n, True, '29', markers, marker1)
