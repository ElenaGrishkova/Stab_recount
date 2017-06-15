# -*- coding: utf-8 -*-
#Программа для объединения файлов из папок отдельных экспериментов в одну сводную таблицу. "Merged_EEH2.xlsx"
# Т.к. состав и название периодов могут быть разные, то они прописаны вручную

from openpyxl import load_workbook
from openpyxl import Workbook

GAME_5STEP_DIR_NAMES={u'08',u'15',u'22',u'29'}
GAME_10STEP_DIR_NAMES={u'11',u'18',u'25',u'01'}
DIR_NAMES=GAME_5STEP_DIR_NAMES | GAME_10STEP_DIR_NAMES
#Имена Листов, которые будут объединены
SHEET_NAMS=[u'Energy V2', u'Energy Vx2+Vy2', u'Energy Vx2', u'Energy Vy2', u'Energy Vz2', u'Energy Vx', u'Energy Vy', u'Energy Vz',
            u'Entropy old', u'Entropy X old', u'Entropy Y old',u'Entropy Z old',u'Entropy new',u'Entropy X',u'Entropy Y', u'Entropy Z',
            u'Hurst V2', u'Hurst Vx2+Vy2', u'Hurst Vx2', u'Hurst Vy2', u'Hurst Vz2', u'Hurst Vx', u'Hurst Vy', u'Hurst Vz',]
#Здесь прописать путь к корневому каталогу на локалоной машине
BASE_DIR_PATH=u'C:\\Users\\egrishkova\\ЭЭ\\Стабилограмма пересчет'
#Имена файлов, которые будут объединяться
SRC_FNAME_BEF='Before_Soc_EEH2.xlsx'
SRC_FNAME_AFT='After_Soc_EEH2.xlsx'
COL_CT_MAX=20
COL_HEIGHT_MAX=11
EMPTY_CELL_VAL=""
IS_READONLY=False
#IS_READONLY=True

#Наименования колонок. Полное объединение! Если здесь не будет хватать каких-то колонок, но они будут присутствовать в исходных файлах
# , то программа будет падать. Если здесь перечислена колонка, которой нет в исходном файле, то в конечный файл добавится пустая именованная колонка
GAME_10STEP_REF_COLNAMS=[u'0', u'ГО', u'ГЗ', u'1',u'1',u'2',u'2',u'3',u'3',u'4',u'4',u'5',u'5',u'6',u'6',u'7',u'7',u'8',u'8',u'9',u'9',u'10',u'10', u'ГО', u'ГЗ',
                         u'1',u'1',u'2',u'2',u'3',u'3',u'4',u'4',u'5',u'5',u'6',u'6',u'7',u'7',u'8',u'8',u'9',u'9',u'10',u'10', u'11',	u'Extr_end', u'ГО', u'ГЗ', u'ГЗ_end'
]
GAME_5STEP_REF_COLNAMS=[u'0', u'ГО', u'ГЗ', u'1',u'1',u'2',u'2',u'3',u'3',u'4',u'4',u'5',u'5',u'1',u'1',u'2',u'2',u'3',u'3',u'4',u'4',u'5',u'5', u'ГО', u'ГЗ',
                        u'1',u'1',u'2',u'2',u'3',u'3',u'4',u'4',u'5',u'5',u'1',u'1',u'2',u'2',u'3',u'3',u'4',u'4',u'5',u'5', u'11',	u'Extr_end', u'ГО', u'ГЗ', u'ГЗ_end']

class Empty_col:
  def __init__(self,ref_nam):
    self.ref_nam=ref_nam

class Empty_cell:
  def __init__(self):
    pass

def fpath_bef(dir_nam):
  return BASE_DIR_PATH+'\\'+dir_nam+'\\'+SRC_FNAME_BEF
def fpath_aft(dir_nam):
  return BASE_DIR_PATH+'\\'+dir_nam+'\\'+SRC_FNAME_AFT

wb_bef_aft_list=[]
for dir in DIR_NAMES:
  wb_bef_aft_list.append((load_workbook(fpath_bef(dir), read_only=IS_READONLY),
                          load_workbook(fpath_aft(dir), read_only=IS_READONLY),
                          dir in GAME_10STEP_DIR_NAMES))

wb_merged = Workbook()
for nam in SHEET_NAMS:
  wb_merged.create_sheet(nam)

for nam in SHEET_NAMS:
  sheet_merged = wb_merged.get_sheet_by_name(nam)

  first_row=True
  for wb_bef,wb_aft,game_10step in wb_bef_aft_list:
    sheet_bef = wb_bef.get_sheet_by_name(nam)
    sheet_aft = wb_aft.get_sheet_by_name(nam)
    humans={row[1].value for row in sheet_bef.iter_rows(min_row=2)} | \
           {row[1].value for row in sheet_aft.iter_rows(min_row=2)}
    humans={h for h in humans if h}
    humans=sorted(humans)
    merged_cols_all={}
    for col_group_id,sheet in {('bef',sheet_bef),('aft',sheet_aft)}:
      src_human_id={}
      id=0
      humans_cur=[ cell.value for cell in list(sheet.columns)[1][1:]]
      for human in humans_cur:
        src_human_id[human]=id
        id+=1
      col_n = 0
      # первый столбец - этоимена людей, но первое значение - не имя, а название столюца с именами
      cols = [[sheet['B1'].value]+humans]
      col_nams = [unicode(i.value) for i in list(sheet.rows)[0][2:]]
      for ref_nam in GAME_10STEP_REF_COLNAMS if game_10step else GAME_5STEP_REF_COLNAMS:
        if col_n<len(col_nams) and ref_nam == col_nams[col_n]:
          new_col=[ref_nam]
          col=list(sheet.columns)[col_n + 2][1:]
          col_n += 1

          for human in humans:
            if human in src_human_id:
              new_col.append(col[src_human_id[human]].value)
            else:
              new_col.append(EMPTY_CELL_VAL)
          cols.append(new_col)
        else:
          cols.append(Empty_col(ref_nam))
      merged_cols_all[col_group_id]=cols

    for row_n in range(len(merged_cols_all['bef'][0])):
      if not row_n and not first_row:
        continue
      first_row=False
      row_merged=[]
      first_col=True
      for cols in [merged_cols_all['bef'],merged_cols_all['aft']]:
        for col_n in range(len(GAME_10STEP_REF_COLNAMS)+1):
          if not col_n and not first_col:
            row_merged.append(EMPTY_CELL_VAL)
            continue
          first_col=False
          if len(cols) == col_n :
            print "col_n" + str(col_n)
          if isinstance(cols[col_n],Empty_col):
            row_merged.append(EMPTY_CELL_VAL if row_n else cols[col_n].ref_nam)
          else:
            row_merged.append(cols[col_n][row_n])

      sheet_merged.append(row_merged)

wb_merged.save('Merged_EEH2.xlsx')







